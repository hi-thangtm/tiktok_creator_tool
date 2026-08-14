"""Excel export service for rows that have valid phone numbers."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from core.contact_utils import (
    build_display_contact,
    format_phone_for_display,
    get_phone_sources,
    join_sources,
    normalize_phone,
    parse_json_dict,
    parse_json_list,
)
from core.constants import status_label


EXPORT_HEADERS = (
    "STT",
    "Mã nhà sáng tạo",
    "Tên hiển thị",
    "Tên người dùng",
    "SĐT chính",
    "Nguồn SĐT",
    "SĐT khác",
    "Email",
    "Nguồn Email",
    "Tiểu sử",
    "Trạng thái",
    "Kiểm tra lần cuối",
)

EXCLUDED_EXPORT_STATUSES = {
    "PENDING",
    "PROCESSING",
    "ERROR",
    "NO_CONTACT",
}

WINDOWS_INVALID_FILENAME_CHARS = set('<>:"/\\|?*')
WINDOWS_RESERVED_FILENAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}


class NoPhoneCreatorsToExport(RuntimeError):
    pass


@dataclass(frozen=True)
class ExportResult:
    path: Path
    exported_count: int


@dataclass(frozen=True)
class ExportPhoneContact:
    main_phone: str
    phone_source: str
    other_phones: str
    normalized_phones: tuple[str, ...]


def build_export_phone_contact(
    row: dict[str, Any],
) -> ExportPhoneContact:
    raw_phone_sources = parse_json_dict(
        row.get("phone_sources")
    )
    official_zalo = normalize_phone(
        row.get("zalo")
    )
    bio_phones = _dedupe_normalized_phones(
        parse_json_list(
            row.get("bio_phones")
        )
    )
    phones = _dedupe_normalized_phones(
        parse_json_list(
            row.get("phones_all")
        )
    )

    for phone in (
        [official_zalo]
        + bio_phones
    ):
        if (
            phone
            and phone not in phones
        ):
            phones.append(phone)

    if not phones:
        return ExportPhoneContact(
            main_phone="",
            phone_source="",
            other_phones="",
            normalized_phones=(),
        )

    if (
        official_zalo
        and official_zalo in phones
    ):
        main_phone = official_zalo
    else:
        main_phone = phones[0]

    other_phones = [
        phone
        for phone in phones
        if phone != main_phone
    ]
    phone_sources = get_phone_sources(
        main_phone,
        raw_phone_sources,
        official_zalo,
        bio_phones,
    )

    return ExportPhoneContact(
        main_phone=format_phone_for_display(main_phone),
        phone_source=join_sources(phone_sources),
        other_phones=", ".join(
            format_phone_for_display(phone)
            for phone in other_phones
        ),
        normalized_phones=tuple(phones),
    )


def creator_is_exportable(
    row: dict[str, Any],
) -> bool:
    status = str(row.get("status") or "").upper()

    if status in EXCLUDED_EXPORT_STATUSES:
        return False

    return bool(
        build_export_phone_contact(row).normalized_phones
    )


def build_export_rows(
    rows: list[dict[str, Any]],
) -> list[list[Any]]:
    result: list[list[Any]] = []

    for row in rows:
        if not creator_is_exportable(row):
            continue

        phone_contact = build_export_phone_contact(row)
        display_contact = build_display_contact(row)
        result.append(
            [
                len(result) + 1,
                str(row.get("creator_id") or ""),
                row.get("nickname") or "",
                row.get("username") or "",
                phone_contact.main_phone,
                phone_contact.phone_source,
                phone_contact.other_phones,
                display_contact["email"],
                display_contact["email_source"],
                row.get("bio") or "",
                status_label(row.get("status")),
                row.get("last_checked_at") or "",
            ]
        )

    return result


def export_creators_to_excel(
    rows: list[dict[str, Any]],
    export_dir: Path,
    filename: str | None = None,
) -> ExportResult:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to export Excel files."
        ) from exc

    export_rows = build_export_rows(rows)

    if not export_rows:
        raise NoPhoneCreatorsToExport(
            "Chưa có nhà sáng tạo nào có số điện thoại để xuất."
        )

    export_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    if filename is None:
        filename = (
            "tiktok_creators_with_phone_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    else:
        filename = sanitize_filename(filename)

    path = _unique_export_path(
        export_dir / filename
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Nhà sáng tạo"
    sheet.append(EXPORT_HEADERS)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:L{len(export_rows) + 1}"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="E5EEF9",
    )
    header_font = Font(
        bold=True,
        color="111827",
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            vertical="center",
        )

    text_columns = {
        2,
        5,
        7,
    }

    for values in export_rows:
        sheet.append(values)
        current_row = sheet.max_row

        for column_index in text_columns:
            cell = sheet.cell(
                row=current_row,
                column=column_index,
            )
            cell.value = str(cell.value or "")
            cell.number_format = "@"

        for column_index in range(
            1,
            len(EXPORT_HEADERS) + 1,
        ):
            sheet.cell(
                row=current_row,
                column=column_index,
            ).alignment = Alignment(
                vertical="top",
                wrap_text=column_index in {10},
            )

    _apply_column_widths(sheet)
    workbook.save(path)

    return ExportResult(
        path=path,
        exported_count=len(export_rows),
    )


def _dedupe_normalized_phones(
    values: list[str],
) -> list[str]:
    result: list[str] = []

    for value in values:
        phone = normalize_phone(value)

        if (
            phone
            and phone not in result
        ):
            result.append(phone)

    return result


def _unique_export_path(
    path: Path,
) -> Path:
    if not path.exists():
        return path

    stem = path.stem
    suffix = path.suffix
    index = 2

    while True:
        candidate = path.with_name(
            f"{stem}_{index}{suffix}"
        )

        if not candidate.exists():
            return candidate

        index += 1


def sanitize_filename(
    filename: str,
    fallback: str = "export.xlsx",
) -> str:
    sanitized = "".join(
        "_"
        if (
            char in WINDOWS_INVALID_FILENAME_CHARS
            or ord(char) < 32
        )
        else char
        for char in filename
    )
    sanitized = sanitized.strip().rstrip(" .")

    if not sanitized:
        return fallback

    path = Path(sanitized)
    stem = path.stem or path.name

    if stem.upper() in WINDOWS_RESERVED_FILENAMES:
        sanitized = f"_{sanitized}"

    return sanitized


def _apply_column_widths(sheet) -> None:
    max_widths = {
        "A": 8,
        "B": 24,
        "C": 28,
        "D": 24,
        "E": 16,
        "F": 14,
        "G": 34,
        "H": 30,
        "I": 16,
        "J": 60,
        "K": 22,
        "L": 22,
    }

    min_widths = {
        "A": 6,
        "B": 22,
        "E": 15,
        "G": 18,
    }

    for column in sheet.columns:
        letter = column[0].column_letter
        width = min_widths.get(letter, 12)
        max_width = max_widths.get(letter, 42)

        for cell in column:
            value = cell.value

            if value is not None:
                width = max(
                    width,
                    min(
                        len(str(value)) + 2,
                        max_width,
                    ),
                )

        sheet.column_dimensions[letter].width = width
