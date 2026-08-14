from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core.database import DatabaseRepository
from services.export_service import export_creators_to_excel


def main() -> int:
    root = Path(
        tempfile.mkdtemp(
            prefix="tiktok_creator_export_smoke_"
        )
    )
    repository = DatabaseRepository(
        root / "creators.db"
    )
    repository.init_database()

    rows = [
        {
            "creator_id": "7494007788183127056",
            "username": "has_phone",
            "nickname": "Has Phone",
            "detail_url": "https://example.test/detail?cid=7494007788183127056",
            "zalo": "+84935850488",
            "official_email": "",
            "bio": "Bio phone 0935 850 488",
            "bio_phones": json.dumps(["0935 850 488"]),
            "bio_emails": "[]",
            "phones_all": json.dumps(["+84935850488"]),
            "emails_all": "[]",
            "phone_sources": json.dumps(
                {
                    "+84935850488": [
                        "Zalo",
                        "Bio",
                    ],
                }
            ),
            "email_sources": "{}",
            "contact_note": "",
            "status": "FOUND_PHONE",
        },
        {
            "creator_id": "7494007788183127057",
            "username": "email_only",
            "nickname": "Email Only",
            "detail_url": "https://example.test/detail?cid=7494007788183127057",
            "zalo": "",
            "official_email": "email@example.com",
            "bio": "Email only",
            "bio_phones": "[]",
            "bio_emails": json.dumps(["email@example.com"]),
            "phones_all": "[]",
            "emails_all": json.dumps(["email@example.com"]),
            "phone_sources": "{}",
            "email_sources": json.dumps(
                {
                    "email@example.com": ["Bio"],
                }
            ),
            "contact_note": "",
            "status": "FOUND_EMAIL",
        },
        {
            "creator_id": "7494007788183127058",
            "username": "pending_phone",
            "nickname": "Pending Phone",
            "detail_url": "https://example.test/detail?cid=7494007788183127058",
            "zalo": "+84974581131",
            "official_email": "",
            "bio": "",
            "bio_phones": "[]",
            "bio_emails": "[]",
            "phones_all": json.dumps(["+84974581131"]),
            "emails_all": "[]",
            "phone_sources": json.dumps(
                {
                    "+84974581131": ["Zalo"],
                }
            ),
            "email_sources": "{}",
            "contact_note": "",
            "status": "PENDING",
        },
    ]

    for row in rows:
        repository.save_creator(row)

    export_result = export_creators_to_excel(
        repository.list_creators_for_phone_export(),
        root / "exports",
        "smoke.xlsx",
    )
    workbook = openpyxl.load_workbook(export_result.path)
    sheet = workbook.active

    print(f"EXPORT_PATH={export_result.path}")
    print(f"EXPORTED_COUNT={export_result.exported_count}")
    print(f"MAX_ROW={sheet.max_row}")
    print(f"CREATOR_ID={sheet['B2'].value}")
    print(f"CREATOR_ID_FORMAT={sheet['B2'].number_format}")
    print(f"MAIN_PHONE={sheet['E2'].value}")
    print(f"MAIN_PHONE_FORMAT={sheet['E2'].number_format}")
    print(f"PHONE_SOURCE={sheet['F2'].value}")

    if export_result.exported_count != 1:
        return 1

    if sheet.max_row != 2:
        return 2

    if sheet["B2"].value != "7494007788183127056":
        return 3

    if (
        sheet["B2"].number_format != "@"
        or sheet["E2"].number_format != "@"
    ):
        return 4

    if sheet["E2"].value != "0935 850 488":
        return 5

    if sheet["F2"].value != "Zalo + Bio":
        return 6

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
